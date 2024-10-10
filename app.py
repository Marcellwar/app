from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify, Response
import psycopg2
import csv
import json
from fpdf import FPDF
import pandas as pd
import io
from werkzeug.security import generate_password_hash, check_password_hash
import xml.etree.ElementTree as ET
from datetime import date
from openpyxl import Workbook

from db import get_db_connection


app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Necesario para manejar sesiones

# Ruta para login
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and user[2] == password:  # Sin bcrypt por simplicidad
            session['user'] = user[1]
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos')

    return render_template('login.html')

# Ruta para dashboard
@app.route('/dashboard')
def dashboard():
    if 'user' in session:
        return render_template('dashboard.html')
    else:
        return redirect(url_for('login'))

# PARTE RELACIONADA A PRODUCTOS.....................................................

# Ruta para listar productos
@app.route('/productos')
def listar_productos():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM producto')
    productos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('productos.html', productos=productos)

# Ruta para agregar producto
@app.route('/productos/agregar', methods=['GET', 'POST'])
def agregar_producto():
    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form['descripcion']
        precio = request.form['precio']
        stock = request.form['stock']
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO producto (nombre, descripcion, precio, stock) VALUES (%s, %s, %s, %s)",
                    (nombre, descripcion, precio, stock))
        conn.commit()
        cur.close()
        conn.close()
        return redirect(url_for('listar_productos'))

    return render_template('agregar_productos.html')

# Ruta para editar producto
@app.route('/productos/editar/<int:id>', methods=['GET', 'POST'])
def editar_producto(id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form['descripcion']
        precio = request.form['precio']
        stock = request.form['stock']
        cur.execute("UPDATE producto SET nombre = %s, descripcion = %s, precio = %s, stock = %s WHERE id = %s",
                    (nombre, descripcion, precio, stock, id))
        conn.commit()
        return redirect(url_for('listar_productos'))
    
    cur.execute('SELECT * FROM producto WHERE id = %s', (id,))
    producto = cur.fetchone()
    cur.close()
    conn.close()

    return render_template('editar_productos.html', producto=producto)

# Ruta para eliminar producto
@app.route('/productos/eliminar/<int:id>')
def eliminar_producto(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM producto WHERE id = %s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('listar_productos'))


# PARTE RELACIONADA A MOVIMIENTOS.....................................................

# Ruta para listar movimientos
@app.route('/movimientos')
def listar_movimientos():
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Actualizamos la consulta para incluir el nombre del usuario
    cur.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre,  -- Obtener el nombre de usuario
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id  -- Unir la tabla de usuarios
    ''')
    
    movimientos = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template('movimientos.html', movimientos=movimientos)

# Ruta para agregar movimiento
@app.route('/movimientos/agregar', methods=['GET', 'POST'])
def agregar_movimiento():
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener la lista de productos
    cur.execute('SELECT * FROM producto')
    productos = cur.fetchall()

    # Obtener la lista de usuarios
    cur.execute('SELECT * FROM usuarios')
    usuarios = cur.fetchall()

    if request.method == 'POST':
        id_producto = request.form['id_producto']  # Obtener el ID del producto seleccionado
        id_usuario = request.form['id_usuario']  # Obtener el ID del usuario seleccionado
        tipo_movimiento = request.form['tipo_movimiento']
        cantidad = request.form['cantidad']
        fecha = request.form['fecha']

        try:
            cur.execute(
                "INSERT INTO movimiento (id_producto, id_usuario, tipo_movimiento, cantidad, fecha) VALUES (%s, %s, %s, %s, %s)",
                (id_producto, id_usuario, tipo_movimiento, cantidad, fecha)
            )
            conn.commit()
            return redirect(url_for('listar_movimientos'))
        except Exception as e:
            conn.rollback()  # Hacer rollback en caso de error
            print(f"Ocurrió un error: {e}")  # Imprimir error para depuración

    cur.close()
    conn.close()
    return render_template('agregar_movimientos.html', productos=productos, usuarios=usuarios)


# Ruta para editar movimiento
@app.route('/movimientos/editar/<int:id>', methods=['GET', 'POST'])
def editar_movimiento(id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener el movimiento actual
    cur.execute('SELECT id, id_producto, tipo_movimiento, cantidad, fecha FROM movimiento WHERE id = %s', (id,))
    movimiento = cur.fetchone()

    # Obtener la lista de productos
    cur.execute('SELECT p.id, p.nombre FROM producto p')
    productos = cur.fetchall()

    if request.method == 'POST':
        # Actualiza el movimiento con los valores proporcionados
        try:
            cur.execute(""" 
                UPDATE movimiento 
                SET id_producto = %s, tipo_movimiento = %s, cantidad = %s, fecha = %s 
                WHERE id = %s
            """, (
                request.form['id_producto'], 
                request.form['tipo_movimiento'], 
                request.form['cantidad'], 
                '2024-09-19 00:00:00',  # Fecha fija como solicitaste
                id
            ))
            conn.commit()
            return redirect(url_for('listar_movimientos'))
        except Exception as e:
            conn.rollback()  # Hacer rollback en caso de error
            print(f"Ocurrió un error: {e}")  # Imprimir error para depuración

    cur.close()
    conn.close()

    return render_template('editar_movimientos.html', movimiento=movimiento, productos=productos)

# Ruta para eliminar movimiento
@app.route('/movimientos/eliminar/<int:id>')
def eliminar_movimiento(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM movimiento WHERE id = %s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('listar_movimientos'))

# Ruta para exportar productos en diferentes formatos
@app.route('/exportar/<formato>')
def exportar_productos(formato):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM producto")
    productos = cur.fetchall()
    column_names = [desc[0] for desc in cur.description]

    if formato == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for producto in productos:
            pdf.cell(200, 10, txt=f"{producto}", ln=True)
        pdf_file = 'productos.pdf'
        pdf.output(pdf_file)
        return send_file(pdf_file, as_attachment=True)

    elif formato == 'xlsx':
        df = pd.DataFrame(productos, columns=column_names)
        excel_file = 'productos.xlsx'
        df.to_excel(excel_file, index=False)
        return send_file(excel_file, as_attachment=True)

    elif formato == 'csv':
        df = pd.DataFrame(productos, columns=column_names)
        csv_file = 'productos.csv'
        df.to_csv(csv_file, index=False)
        return send_file(csv_file, as_attachment=True)

    elif formato == 'xml':
        root = ET.Element("Productos")
        for producto in productos:
            producto_element = ET.SubElement(root, "Producto")
            for i, col in enumerate(column_names):
                ET.SubElement(producto_element, col).text = str(producto[i])
        tree = ET.ElementTree(root)
        xml_file = 'productos.xml'
        tree.write(xml_file)
        return send_file(xml_file, as_attachment=True)

    elif formato == 'json':
        data = []
        for prod in productos:
            producto_dict = {
                "ID": prod[0],
                "Nombre": prod[1],
                "Descripcion": prod[2],
                "Precio": prod[3],
                "Stock": prod[4]
            }
            data.append(producto_dict)

        json_output = json.dumps(data, indent=4)
        return send_file(io.BytesIO(json_output.encode()), as_attachment=True, download_name='productos.json', mimetype='application/json')

    return redirect(url_for('listar_productos'))


# Ruta para exportar movimientos en diferentes formatos
@app.route('/exportar/movimientos_/<formato>')
def exportar_movimientos(formato):
    conn = get_db_connection()
    cur = conn.cursor()

    # Nueva consulta SQL uniendo movimiento, producto y usuarios
    cur.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre,  -- Obtener el nombre de usuario
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id  -- Unir la tabla de usuarios
    ''')
    
    movimientos = cur.fetchall()
    column_names = ['ID', 'Producto', 'Usuario', 'Tipo de Movimiento', 'Cantidad', 'Fecha']

    if formato == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for movimiento in movimientos:
            pdf.cell(200, 10, txt=f"{movimiento}", ln=True)
        pdf_file = 'movimientos.pdf'
        pdf.output(pdf_file)
        return send_file(pdf_file, as_attachment=True)

    elif formato == 'xlsx':
        df = pd.DataFrame(movimientos, columns=column_names)
        excel_file = 'movimientos.xlsx'
        df.to_excel(excel_file, index=False)
        return send_file(excel_file, as_attachment=True)

    elif formato == 'csv':
        df = pd.DataFrame(movimientos, columns=column_names)
        csv_file = 'movimientos.csv'
        df.to_csv(csv_file, index=False)
        return send_file(csv_file, as_attachment=True)

    elif formato == 'xml':
        root = ET.Element("Movimientos")
        for movimiento in movimientos:
            movimiento_element = ET.SubElement(root, "Movimiento")
            for i, col in enumerate(column_names):
                ET.SubElement(movimiento_element, col).text = str(movimiento[i])
        tree = ET.ElementTree(root)
        xml_file = 'movimientos.xml'
        tree.write(xml_file)
        return send_file(xml_file, as_attachment=True)

    elif formato == 'json':
        data = []
        for mov in movimientos:
            movimiento_dict = {
                "ID": mov[0],
                "Producto": mov[1],
                "Usuario": mov[2],
                "Tipo de Movimiento": "Entrada" if mov[3] == 1 else "Salida",
                "Cantidad": mov[4],
                "Fecha": mov[5].strftime("%Y-%m-%d %H:%M:%S") if mov[5] else None
            }
            data.append(movimiento_dict)

        json_output = json.dumps(data, indent=4)
        return send_file(io.BytesIO(json_output.encode()), as_attachment=True, download_name='movimientos.json', mimetype='application/json')

    return redirect(url_for('listar_movimientos'))



# Ruta para generar reportes 
@app.route('/reportes')
def reportes():
    return render_template('reportes.html')


# Cerrar sesión
@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('id', None)
    session.pop('username', None)
    flash('Has cerrado sesión correctamente.', 'success')
    return redirect(url_for('login'))

@app.route('/buscar_productos')
def buscar_productos():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM producto WHERE nombre ILIKE %s OR id::text ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    # Convertir a un formato JSON adecuado
    result = []
    for producto in productos:
        result.append({
            'id': producto[0],
            'nombre': producto[1],
            'descripcion': producto[2],
            'precio': producto[3],
            'stock': producto[4]
        })

    return jsonify(result)

@app.route('/movimientos_producto')
def movimientos_producto():
    producto_id = request.args.get('id')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT m.id, m.id_producto, u.username AS usuario, m.tipo_movimiento, m.cantidad, m.fecha 
        FROM movimiento m
        JOIN usuarios u ON m.id_usuario = u.id
        WHERE m.id_producto = %s
    """, (producto_id,))
    movimientos = cursor.fetchall()

    # Convertir a un formato JSON adecuado
    result = []
    for movimiento in movimientos:
        result.append({
            'id': movimiento[0],
            'id_producto': movimiento[1],
            'usuario': movimiento[2],
            'tipo_movimiento': movimiento[3],
            'cantidad': movimiento[4],
            'fecha': movimiento[5].isoformat()  # Convertir a formato ISO para JSON
        })

    return jsonify(result)

@app.route('/export/csv')
def exportar_csv():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM producto WHERE nombre ILIKE %s OR id::text ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    # Crear CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Nombre', 'Descripción', 'Precio', 'Stock'])
    for producto in productos:
        writer.writerow(producto)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), as_attachment=True, download_name='productos.csv', mimetype='text/csv')


@app.route('/export/xlsx')
def exportar_xlsx():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM producto WHERE nombre ILIKE %s OR id::text ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"

    # Escribir encabezados
    sheet.append(['ID', 'Nombre', 'Descripción', 'Precio', 'Stock'])

    for producto in productos:
        sheet.append(producto)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='productos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/pdf')
def exportar_pdf():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM producto WHERE nombre ILIKE %s OR id::text ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Reporte de Productos", ln=True, align='C')

    for producto in productos:
        pdf.cell(200, 10, txt=f"ID: {producto[0]} - Nombre: {producto[1]} - Descripción: {producto[2]} - Precio: {producto[3]} - Stock: {producto[4]}", ln=True)
    pdf_file = 'productos.pdf'
    pdf.output(pdf_file)
    return send_file(pdf_file, as_attachment=True)

@app.route('/export/json')
def export_json():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM producto WHERE nombre ILIKE %s OR id::text ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    result = [{'id': p[0], 'nombre': p[1], 'descripcion': p[2], 'precio': p[3], 'stock': p[4]} for p in productos]

    response = json.dumps(result)
    return Response(response, mimetype='application/json', headers={"Content-Disposition": "attachment;filename=productos.json"})

@app.route('/export/xml')
def exportar_xml():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM producto WHERE nombre ILIKE %s OR id::text ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    root = ET.Element("productos")
    for producto in productos:
        prod_elem = ET.SubElement(root, "producto")
        ET.SubElement(prod_elem, "id").text = str(producto[0])
        ET.SubElement(prod_elem, "nombre").text = producto[1]
        ET.SubElement(prod_elem, "descripcion").text = producto[2] if producto[2] else ""
        ET.SubElement(prod_elem, "precio").text = str(producto[3]) if producto[3] else "0.0"
        ET.SubElement(prod_elem, "stock").text = str(producto[4])

    output = io.BytesIO()
    tree = ET.ElementTree(root)
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='productos.xml', mimetype='application/xml')

@app.route('/exportar/movimientos/csv')
def exportar_movimientos_csv():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre, 
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id 
        WHERE producto.nombre ILIKE %s OR usuarios.username ILIKE %s OR movimiento.id::text ILIKE %s
    ''', (f'%{term}%', f'%{term}%', f'%{term}%'))
    
    movimientos = cursor.fetchall()

    # Crear CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Producto', 'Usuario', 'Tipo de Movimiento', 'Cantidad', 'Fecha'])
    for movimiento in movimientos:
        writer.writerow(movimiento)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), as_attachment=True, download_name='movimientos.csv', mimetype='text/csv')

@app.route('/exportar/movimientos/xlsx')
def exportar_movimientos_xlsx():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre, 
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id 
        WHERE producto.nombre ILIKE %s OR usuarios.username ILIKE %s OR movimiento.id::text ILIKE %s
    ''', (f'%{term}%', f'%{term}%', f'%{term}%'))
    
    movimientos = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimientos"

    # Escribir encabezados
    sheet.append(['ID', 'Producto', 'Usuario', 'Tipo de Movimiento', 'Cantidad', 'Fecha'])

    for movimiento in movimientos:
        sheet.append(movimiento)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='movimientos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar/movimientos/pdf')
def exportar_movimientos_pdf():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre, 
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id 
        WHERE producto.nombre ILIKE %s OR usuarios.username ILIKE %s OR movimiento.id::text ILIKE %s
    ''', (f'%{term}%', f'%{term}%', f'%{term}%'))
    
    movimientos = cursor.fetchall()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Reporte de Movimientos", ln=True, align='C')

    for movimiento in movimientos:
        pdf.cell(200, 10, txt=f"ID: {movimiento[0]} - Producto: {movimiento[1]} - Usuario: {movimiento[2]} - Tipo: {'Entrada' if movimiento[3] == 1 else 'Salida'} - Cantidad: {movimiento[4]} - Fecha: {movimiento[5]}", ln=True)

    pdf_file = 'movimientos.pdf'
    pdf.output(pdf_file)
    return send_file(pdf_file, as_attachment=True)

@app.route('/exportar/movimientos/json')
def exportar_movimientos_json():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre, 
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id 
        WHERE producto.nombre ILIKE %s OR usuarios.username ILIKE %s OR movimiento.id::text ILIKE %s
    ''', (f'%{term}%', f'%{term}%', f'%{term}%'))
    
    movimientos = cursor.fetchall()

    result = []
    for mov in movimientos:
        result.append({
            'id': mov[0],
            'producto': mov[1],
            'usuario': mov[2],
            'tipo_movimiento': 'Entrada' if mov[3] == 1 else 'Salida',
            'cantidad': mov[4],
            'fecha': mov[5].strftime("%Y-%m-%d %H:%M:%S") if mov[5] else None
        })

    response = json.dumps(result)
    return Response(response, mimetype='application/json', headers={"Content-Disposition": "attachment;filename=movimientos.json"})


@app.route('/exportar/movimientos/xml')
def exportar_movimientos_xml():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 
            movimiento.id, 
            producto.nombre AS producto_nombre, 
            usuarios.username AS usuario_nombre, 
            movimiento.tipo_movimiento, 
            movimiento.cantidad, 
            movimiento.fecha 
        FROM movimiento 
        JOIN producto ON movimiento.id_producto = producto.id 
        JOIN usuarios ON movimiento.id_usuario = usuarios.id 
        WHERE producto.nombre ILIKE %s OR usuarios.username ILIKE %s OR movimiento.id::text ILIKE %s
    ''', (f'%{term}%', f'%{term}%', f'%{term}%'))
    
    movimientos = cursor.fetchall()

    root = ET.Element("movimientos")
    for movimiento in movimientos:
        mov_elem = ET.SubElement(root, "movimiento")
        ET.SubElement(mov_elem, "id").text = str(movimiento[0])
        ET.SubElement(mov_elem, "producto").text = movimiento[1]
        ET.SubElement(mov_elem, "usuario").text = movimiento[2]
        ET.SubElement(mov_elem, "tipo_movimiento").text = 'Entrada' if movimiento[3] == 1 else 'Salida'
        ET.SubElement(mov_elem, "cantidad").text = str(movimiento[4])
        ET.SubElement(mov_elem, "fecha").text = movimiento[5].strftime("%Y-%m-%d %H:%M:%S") if movimiento[5] else None

    output = io.BytesIO()
    tree = ET.ElementTree(root)
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='movimientos.xml', mimetype='application/xml')


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)
